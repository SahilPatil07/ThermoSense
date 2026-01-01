import re
import difflib
import logging
from pathlib import Path
from typing import List, Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

logger = logging.getLogger(__name__)

def list_sheetnames(file_path: str) -> List[str]:
    wb = load_workbook(filename=str(file_path), data_only=True, read_only=True, keep_vba=True)
    try:
        return list(wb.sheetnames)
    finally:
        try: wb.close()
        except Exception: pass

def _normalize_label(s: str) -> str:
    if s is None:
        return ""
    s = s.replace("\xa0", " ")
    s = s.strip().lower()
    s = re.sub(r"[_\-]+", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s

def resolve_sheet_name(file_path: str, requested: str) -> Optional[str]:
    if not requested:
        return None
    sheets = list_sheetnames(file_path)
    if requested in sheets:
        return requested
    for s in sheets:
        if s.lower() == requested.lower():
            return s
    req_n = _normalize_label(requested)
    for s in sheets:
        if _normalize_label(s) == req_n:
            return s
    best_name = None
    best_score = 0.0
    for s in sheets:
        score = difflib.SequenceMatcher(None, req_n, _normalize_label(s)).ratio()
        if score > best_score:
            best_score = score
            best_name = s
    if best_score >= 0.65:
        return best_name
    return None

def _detect_header_row_fast(values, scan_rows: int = 12) -> int:
    n = min(scan_rows, len(values))
    best_idx, best_score = 0, -1
    for i in range(n):
        row = values[i]
        score = sum(1 for v in row if (v is not None and str(v).strip() != ""))
        if score > best_score:
            best_score = score
            best_idx = i
    return best_idx

def _clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.dropna(how="all")
    if not df.empty:
        df = df.loc[:, ~df.isna().all()]
    
    # Deduplicate columns
    new_cols = []
    seen = {}
    for idx, c in enumerate(df.columns):
        c_str = str(c).strip() if c is not None else ""
        
        # Handle empty or unnamed columns
        if c_str == "" or c_str.lower().startswith("unnamed:"):
            c_str = f"Col_{idx+1}"
            
        # Deduplicate
        if c_str in seen:
            seen[c_str] += 1
            c_str = f"{c_str}_{seen[c_str]}"
        else:
            seen[c_str] = 0
            
        new_cols.append(c_str)
        
    df.columns = new_cols
    return df

def _values_from_range(ws, a1_range: str):
    """
    Extract values from a range safely.
    """
    try:
        if not a1_range or ":" not in a1_range:
            logger.warning(f"Invalid range format: {a1_range}. Defaulting to A1:Z100")
            a1_range = "A1:Z100"
            
        min_col, min_row, max_col, max_row = range_boundaries(a1_range)
        
        # Clamp values to worksheet limits
        max_row = min(max_row, ws.max_row or 10000)
        max_col = min(max_col, ws.max_column or 1000)
        
        out = []
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
            out.append(list(row))
        return out
    except Exception as e:
        logger.error(f"Error iterating over range {a1_range}: {e}")
        # Fallback
        try:
            out = []
            for r in range(1, 100):
                row = [ws.cell(row=r, column=c).value for c in range(1, 20)]
                if any(v is not None for v in row):
                    out.append(row)
                else:
                    break
            return out
        except:
            return []

def _used_range(ws) -> str:
    """
    Get the used range of a worksheet.
    Handles edge cases where calculate_dimension() fails.
    """
    try:
        dim = ws.calculate_dimension()
        if ":" not in dim:
            # Single cell or malformed range, fallback
            max_row = ws.max_row or 1
            max_col = ws.max_column or 1
            from openpyxl.utils import get_column_letter
            return f"A1:{get_column_letter(max_col)}{max_row}"
        return dim
    except Exception as e:
        # Fallback when calculate_dimension() fails (e.g., invalid cell references)
        logger.warning(f"calculate_dimension() failed: {e}. Using max_row/max_col fallback.")
        try:
            max_row = ws.max_row or 1
            max_col = ws.max_column or 1
            from openpyxl.utils import get_column_letter
            fallback_range = f"A1:{get_column_letter(max_col)}{max_row}"
            logger.info(f"Fallback range: {fallback_range}")
            return fallback_range
        except Exception as fallback_error:
            # Ultimate fallback - assume reasonable size
            logger.error(f"Even fallback failed: {fallback_error}. Using default A1:Z1000")
            return "A1:Z1000"

def read_excel_table(
    file_path: str,
    sheet_name: str,
    a1_range: Optional[str] = None,
    header_row: Optional[int] = None,
) -> pd.DataFrame:
    from backend.settings import EXCEL_HEADER_SCAN_ROWS
    file_path = str(file_path)
    resolved = resolve_sheet_name(file_path, sheet_name)
    sheets = list_sheetnames(file_path)
    if not resolved:
        raise ValueError(f"Worksheet '{sheet_name}' not found. Available: {sheets}")

    wb = load_workbook(filename=file_path, data_only=True, read_only=True, keep_vba=True)
    try:
        ws = wb[resolved]
        if a1_range:
            values = _values_from_range(ws, a1_range)
        else:
            used = _used_range(ws)
            values = _values_from_range(ws, used)
    finally:
        try: wb.close()
        except Exception: pass

    if not values:
        return pd.DataFrame()

    if header_row is not None and header_row >= 1:
        hdr_idx = header_row - 1
    else:
        hdr_idx = _detect_header_row_fast(values, scan_rows=EXCEL_HEADER_SCAN_ROWS)

    hdr_idx = max(0, min(hdr_idx, len(values) - 1))
    header = [("" if v is None else str(v)) for v in values[hdr_idx]]
    rows = values[hdr_idx + 1 :]
    df = pd.DataFrame(rows, columns=header)
    df = _clean_dataframe(df)
    return df